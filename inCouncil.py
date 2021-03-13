import openpyxl


def createWorkSheet():
    file = 'HackLboro_-_Team_2_-_Database.xlsx'
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook['Councils']
    return worksheet


def returnCouncilList():
    worksheet = createWorkSheet()
    councils = []
    for r in range(2, worksheet.max_row + 1):
        councilItem = worksheet.cell(row=r, column=1).value
        councils.append(councilItem)
    return councils


def returnMatType():
    file = 'HackLboro_-_Team_2_-_Database.xlsx'
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook['Materials']
    types = []
    for r in range(2, worksheet.max_row + 1):
        typeItem = worksheet.cell(row=r, column=1).value
        if (typeItem not in types):
            types.append(typeItem)
    return types


def returnMaterialList(typeMat):
    file = 'HackLboro_-_Team_2_-_Database.xlsx'
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook['Materials']
    materials = []
    for r in range(2, worksheet.max_row + 1):
        typeItem = worksheet.cell(row=r, column=1).value
        materialItem = worksheet.cell(row=r, column=2).value
        if(typeItem == typeMat):
            materials.append(materialItem)
    return materials


def performLookUp(council, material):
    worksheet = createWorkSheet()
    gotIt = ""
    for r in range(2, worksheet.max_row + 1):
        councilItem = worksheet.cell(row=r, column=1).value
        if councilItem == council:
            for c in range(2, worksheet.max_column + 1):
                materialItem = worksheet.cell(row=1, column=c).value
                if materialItem == material:
                    gotIt = worksheet.cell(row=r, column=c).value
                    break
            break
    return gotIt


if __name__ == "__main__":
    print(createWorkSheet())
    print(returnCouncilList())
    print(returnMatType())
    print(returnMaterialList("Automotive"))
    print(performLookUp('Blackburn with Darwen Borough Council', 'Tyres'))
    print(performLookUp('Arun District Council', 'Asbestos'))
